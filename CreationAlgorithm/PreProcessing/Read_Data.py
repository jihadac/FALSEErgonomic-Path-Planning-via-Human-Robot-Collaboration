# Script to extract joints positions from Excel files, obtained with XSens
from openpyxl import load_workbook
import numpy as np
import copy

def extract_joint_posisitions_for_fabrik():
    
    joint_positions = extract_joints_positions()
    rotated_joint_positions = rotate_joint_positions(joint_positions)
    adapted_joint_positions = account_for_height_table(rotated_joint_positions)
    joint_positions_fabrik = rotate_for_fabrik(adapted_joint_positions)
    
    return joint_positions_fabrik
    

def extract_joints_positions():
    
    file = 'C:/Users\jihad\Documents\VUB\Master 2\Master Thesis\Code\XSens/Exports/XSens-FrontTable#Jihad.xlsx' # Path to Excel file
    
    wb = load_workbook(file,data_only=True)           # Load Excel file
    ws = wb['Segment Position']                       # Desired worksheet
    
    counter = 0
    desired_columns = ['L3 x','L3 y','L3 z',
                       'Right Upper Arm x','Right Upper Arm y','Right Upper Arm z',
                       'Right Forearm x','Right Forearm y','Right Forearm z',
                       'Right Hand x','Right Hand y','Right Hand z',
                       'Right Finger x','Right Finger y','Right Finger z',
                       'Left Upper Arm x','Left Upper Arm y','Left Upper Arm z',
                       'Left Forearm x','Left Forearm y','Left Forearm z',
                       'Left Hand x','Left Hand y','Left Hand z',
                       'Left Finger x','Left Finger y','Left Finger z',
                       'Right Upper Leg x','Right Upper Leg y','Right Upper Leg z',
                       'Left Upper Leg x','Left Upper Leg y','Left Upper Leg z',
                       'Right Lower Leg x','Right Lower Leg y','Right Lower Leg z',
                       'Right Foot x','Right Foot y','Right Foot z',
                       'Right Toe x','Right Toe y','Right Toe z',
                       'Left Lower Leg x','Left Lower Leg y','Left Lower Leg z',
                       'Left Foot x','Left Foot y','Left Foot z',
                       'Left Toe x','Left Toe y','Left Toe z',
                       'Neck x','Neck y','Neck z',
                       'Head x','Head y','Head z']
    
    last_row = ws.max_row # Extract values of last row
    
    extracted_values = []
    list_joint = []
    
    for name in desired_columns:
        column_name = get_column_name(ws,name) # Get the letter (name) of the corresponding column
        cell_value = ws[column_name + str(last_row)].value
        list_joint.append(cell_value)
        counter += 1
        if counter == 3:
            extracted_values.append(list_joint)
            counter = 0
            list_joint = []
   
    #print(extracted_values)
            
    return extracted_values

def get_column_name(ws, element):
    for cell in ws[1]: # Iterate through the cells in the first row
        if cell.value == element:
            column_name = cell.column_letter
            return column_name
        
def rotate_joint_positions(extracted_values):
    
    R = [[-1,0,0],[0,-1,0],[0,0,1]] # Rotation around z-axis over 180°
    
    rotated_joints = []
    
    for joint in extracted_values:
        rotated_joint = np.matmul(R,joint).tolist()
        rotated_joints.append(rotated_joint)
    
    return rotated_joints

def account_for_height_table(joints): # The joint coordinates should be computed with respect to the robot's root frame. The robot is placed on a table with a certain height. 
   
    height_table = 0.845 #0.955938497  # Height is in z-direction
    
    # To account for the height of the table, the z-coordinate of all the joints should be subtracted by the height of the table.
    # This is because the joints' positions are now defined with respect to the ground. But they should be defined with respect to the frame of the root of the robot.
    joints_with_height = copy.deepcopy(joints)  # Create a copy of the joints list
    for joint in joints_with_height:
        joint[2] = joint[2]-height_table # Index 2, because third element (counting starts from zero) of joint (= list) corresponds to z-coordinate
    
    return joints_with_height

def rotate_for_fabrik(joints):
    R = [[0,-1,0],[1,0,0],[0,0,1]] # Rotation around z-axis over 90°
    
    rotated_joints = []
    
    for joint in joints:
        rotated_joint = np.matmul(R,joint).tolist()
        rotated_joints.append(rotated_joint)
        
    return rotated_joints

#a = extract_joints_positions()
#b = rotate_joint_positions(a)
#c = account_for_height_table(b)
#final = rotate_for_fabrik(c)
