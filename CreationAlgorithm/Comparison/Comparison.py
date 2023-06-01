import Candidate_Points as points
import Score_Candidate_Points as scores
import Graph as graph
import A_Star_Comparison as a_star
import Dijkstra as dijkstra
import Drawing_Graph as circle_path
import Metrics as metrics
import random
import math
from openpyxl import load_workbook
from memory_profiler import profile
import Results as draw_results



# Parameters ellipsoid
a = 0.855
b = 0.855
c = 0.857

step = 0.15	                   # Distance between the candidate points
height_circle = 0.333	       # Height centre point circle with respect to origin
distance_circle = 0.33         # Distance in y direction with respect to origin --> determines how far circle is from human

# Compute the x- and z-coordinates of the candidate points inside the circle
tuples, a_new, c_new = points.candidatePoints(a,b,c,distance_circle,height_circle,step)	
length = len(tuples)

# Compute the REBA score of each candidate point
    # colors is a list of colors, to every point a color is associated depending on the corresponding REBA score
    # REBA_scores_list is a list containing the REBA score of every candidate point
    # results_fabrik_list is a list containing the result of FABRIK for each candidate point
colors, REBA_scores_list, results_fabrik_list, error = scores.computeScores(tuples,distance_circle)

# Create a dictionary to associate each tuple to the corresponding REBA score
REBA_tuple_dictionary = dict() 
for i in range(length):
	REBA_tuple_dictionary.update({tuples[i]:REBA_scores_list[i]}) # Dictionary that stores for every tuple (point) the corresponding REBA score

# Compute the adjacancy matrix necessary to create the graph
    # M represents the adjacency matrix
    # tuples is a list of tuples, representing the x and z coordinate of each candidate point (x,z)
M = graph.createAdjacencyMatrix(tuples, step, length, REBA_scores_list)

# Create graph
G = graph.createGraph(M, tuples, length)

# Create a list of random starting nodes
	# Repetition defines the number of paths will be generated for different random initial nodes
	# number_of_paths is a counter to keep track of the number of paths already found
	# directory represents the directory where the paths will be saved
repetition = length
list_starting_nodes = random.sample(tuples,repetition)
path_number = 1
# ws1: list_starting_nodes = [(0.1113, 1.0), (0.2613, 1.0), (-0.1887, 0.85), (-0.3387, 1.0), (0.2613, 0.7), (-0.6387, 0.55), (0.1113, 0.1), (-0.3387, 0.25), (0.2613, 0.55), (-0.0387, 0.1), (0.1113, 0.7), (-0.0387, -0.2), (0.7113, 0.4), (0.7113, 0.25), (-0.4887, -0.2), (-0.0387, 0.25), (0.7113, 0.55), (-0.1887, 0.7), (0.2613, 0.85), (0.4113, -0.05), (-0.1887, 0.55), (-0.3387, 0.1), (-0.4887, 0.25), (0.1113, 0.55), (-0.4887, 0.1), (-0.1887, -0.05), (0.4113, 0.25), (-0.4887, 0.7), (0.2613, -0.05), (0.1113, 0.85), (-0.6387, 0.4), (-0.4887, 0.4), (0.1113, 0.25), (-0.4887, 0.85), (-0.4887, -0.05), (-0.3387, -0.2), (0.5613, -0.05), (-0.6387, -0.05), (0.4113, 0.85), (-0.1887, 0.4), (-0.0387, 0.55), (0.4113, 0.7), (0.5613, 0.25), (-0.3387, 0.85), (-0.4887, 0.55), (-0.6387, 0.1), (0.4113, 0.4), (0.1113, 0.4), (0.2613, 0.4), (0.4113, -0.2), (0.5613, 0.4), (0.4113, 0.1), (-0.1887, 0.25), (-0.3387, -0.05), (0.5613, 0.1), (-0.0387, 0.4), (0.4113, 1.0), (0.5613, 0.55), (0.5613, 0.7), (-0.0387, 0.7), (-0.3387, 0.7), (-0.3387, 0.4), (-0.0387, 0.85), (0.2613, 0.25), (-0.6387, 0.25), (-0.6387, 0.7), (-0.0387, 1.0), (0.2613, -0.2), (0.1113, -0.2), (0.7113, 0.1), (-0.1887, 1.0), (-0.3387, 0.55), (0.5613, -0.2), (-0.1887, -0.2), (0.4113, 0.55), (0.5613, 0.85), (-0.1887, 0.1), (0.2613, 0.1), (0.1113, -0.05), (-0.0387, -0.05)]
# Find the goal node, which is necessary to compute the most ergonomic path
min_score = min(REBA_tuple_dictionary.values())
possible_goal_nodes = [node for node, score in REBA_tuple_dictionary.items() if score == min_score]

draw_results.plotCircle(colors, a_new, b, c_new, REBA_tuple_dictionary, length, height_circle)

# Load workbook
workbook = load_workbook('C:/Users\jihad\Documents\VUB\Master 2\Master Thesis/Comparison.xlsx')
worksheet = workbook['Workspace3']



for start_node in list_starting_nodes: 

    # If multiple candidate points have the same lowest REBA score, define goal_node as the one closest to start_node
    goal_node = min(possible_goal_nodes, key = lambda t: math.dist(t, start_node))

    # Execute the A* algorithm
    path_a_star1, runtime1, expanded_nodes1 = a_star.A_Star_Algorithm_Comparison(G, start_node, goal_node, M, REBA_tuple_dictionary, tuples,1)
    path_a_star2, runtime2, expanded_nodes2 = a_star.A_Star_Algorithm_Comparison(G, start_node, goal_node, M, REBA_tuple_dictionary, tuples,2)
    path_a_star3, runtime3, expanded_nodes3 = a_star.A_Star_Algorithm_Comparison(G, start_node, goal_node, M, REBA_tuple_dictionary, tuples,3)

    # Execute Dijkstra's algorithm
    path_dijkstra, runtime_d, expanded_nodes_d = dijkstra.Dijkstras_Algorithm(G, start_node, goal_node, M, tuples)

    # Length paths
    length_path1 = metrics.compute_length_path(path_a_star1)
    length_path2 = metrics.compute_length_path(path_a_star2)
    length_path3 = metrics.compute_length_path(path_a_star3)
    length_path_d = metrics.compute_length_path(path_dijkstra)
    
    # Score paths
    score_path1 = metrics.compute_total_score_path(path_a_star1, REBA_tuple_dictionary)
    score_path2 = metrics.compute_total_score_path(path_a_star2, REBA_tuple_dictionary)
    score_path3 = metrics.compute_total_score_path(path_a_star3, REBA_tuple_dictionary)
    score_path_d = metrics.compute_total_score_path(path_dijkstra, REBA_tuple_dictionary)

    # Number of nodes
    number_nodes1 = metrics.number_of_nodes_path(path_a_star1)
    number_nodes2 = metrics.number_of_nodes_path(path_a_star2)
    number_nodes3 = metrics.number_of_nodes_path(path_a_star3)
    number_nodes_d = metrics.number_of_nodes_path(path_dijkstra)
    
    start_node = str(start_node)
    
    variables = [path_number, start_node, 
                  runtime_d,runtime1,runtime2,runtime3,
                  expanded_nodes_d,expanded_nodes1,expanded_nodes2,expanded_nodes3,
                  length_path_d,length_path1,length_path2,length_path3,
                  score_path_d,score_path1,score_path2,score_path3,
                  number_nodes_d,number_nodes1,number_nodes2,number_nodes3]
    
    # Write the variables to consecutive cells in a row
    for i in range(len(variables)):
        cell = worksheet.cell(row=path_number+1, column=i+1)
        cell.value = variables[i]

    workbook.save('C:/Users\jihad\Documents\VUB\Master 2\Master Thesis/Comparison.xlsx')
    
    path_number += 1
    
    # print(path_a_star1,expanded_nodes1,runtime1)
    # print(path_a_star2,expanded_nodes2,runtime2)
    # print(path_a_star3,expanded_nodes3,runtime3)
    # print(path_dijkstra,expanded_nodes_d,runtime_d)
    
    # print('Length path',length_path1,length_path2,length_path3,length_path_d)
    # print('Score path',score_path1,score_path2,score_path3,score_path_d)
    # print('Number of nodes',number_nodes1,number_nodes2,number_nodes3,number_nodes_d)

    # # Draw path on circle
    # circle_path.drawPath(path_a_star1, REBA_tuple_dictionary, colors, a_new, b, c_new, length, height_circle)
    # circle_path.drawPath(path_a_star2, REBA_tuple_dictionary, colors, a_new, b, c_new, length, height_circle)
    # circle_path.drawPath(path_a_star2, REBA_tuple_dictionary, colors, a_new, b, c_new, length, height_circle)
    # circle_path.drawPath(path_dijkstra, REBA_tuple_dictionary, colors, a_new, b, c_new, length, height_circle)
    