from openpyxl import load_workbook

file = 'C:/Users/jihad/Documents/VUB/Master 2/Master Thesis/Code/XSens/OnlyAngles/XSens-Distance3_Part6#Jihad.xlsx' # Path to Excel file
file_destination = 'C:/Users/jihad/Documents/VUB/Master 2/Master Thesis/Code/XSens/Joint_Angles_XSens.xlsx'
# Load data Excel file
wb = load_workbook(file,data_only=True)           
worksheets_names = ['Joint Angles ZXY','Ergonomic Joint Angles ZXY']
worksheets = []

for sheet_name in worksheets_names:
    ws = wb[sheet_name]
    worksheets.append(ws)
    
# Load the destination workbook
destination_wb = load_workbook(file_destination,data_only=True)
destination_ws = destination_wb['Distance3']

# Extract values from specific rows
start_row = 54
end_row = 68

# Extract the values from the first column within the specified range of rows
frames_motion = [destination_ws.cell(row=row, column=1).value for row in range(start_row, end_row + 1)]

column_names_ergonomic = ['D','B','C','S','Q','R','J','G','H','E']
column_names_normal = ['AW','BI','AB','AN','AE','AQ','AC','AO','AD','AP']

own_ergonomic = ['B','C','D','E','F','G','J','K','L','M']
own_normal = ['H','I','P','Q','R','S','T','U','V','W']

length_ergonomic = len(column_names_ergonomic)
length_normal = len(column_names_normal)
counter = start_row

worksheet_normal = worksheets[0]
worksheet_ergo = worksheets[1]

for frame in frames_motion:
    for i in range(length_ergonomic):    
        cell_value = worksheet_ergo[column_names_ergonomic[i] + str(frame+2)].value # +2 because frame 1 corresponds to row 3 in Excel file
        destination_ws[own_ergonomic[i]+str(counter)] = cell_value
    for i in range(length_normal):
        cell_value = worksheet_normal[column_names_normal[i] + str(frame+2)].value
        destination_ws[own_normal[i]+str(counter)] = cell_value
    
    destination_wb.save(file_destination)
    counter = counter + 1


        